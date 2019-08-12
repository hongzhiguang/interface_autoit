from openpyxl import load_workbook
import os
import traceback
import time

class ParseExcel(object):
    """"此类主要封装常用的操作excel文件的方法"""

    def __init__(self,excel_file_path):
        if not os.path.exists(excel_file_path):
            self.wb = None
        self.excel_file_path = excel_file_path
        self.wb = load_workbook(excel_file_path)
        # 初始化的时候，默认ws为第一个sheet
        self.ws = self.set_sheet_by_name(self.wb.sheetnames[0])

    def create_new_sheet(self, name):
        # 新建一个sheet
        self.wb.create_sheet(name)
        self._save()

    def get_current_sheet(self):
        # 获取当前的sheetname
        if self.ws is not None:
            return self.ws.title
        return None

    def set_sheet_by_name(self, name):
        # 设置ws
        if name in self.wb.sheetnames:
            self.ws = self.wb[name]
            return self.ws
        self.ws = None
        return self.ws

    def get_min_row(self):
        # 获取表格中有数据的最小行数
        try:
            return self.ws.min_row
        except:
            traceback.print_exc()
            return None

    def get_max_row(self):
        # 获取表格中有数据的最大行数
        try:
            return self.ws.max_row
        except:
            traceback.print_exc()
            return None

    def get_row(self, row_no):
        # self.ws.rows返回包含所有行的一个迭代器
        if not isinstance(row_no, int):
            return None
        try:
            # for i in self.ws.rows:
            #    print(i)
            return list(self.ws.rows)[row_no - 1]
        except:
            traceback.print_exc()

    def get_col(self, col_no):
        # self.ws.columns返回包含所有列的一个迭代器
        if not isinstance(col_no, int):
            return None
        try:
            return list(self.ws.columns)[col_no - 1]
        except:
            traceback.print_exc()

    def get_cell_value(self, row_no, col_no):
        """参数行号和列表从1开始表示第一行"""
        if (not isinstance(row_no, int)) or (not isinstance(col_no, int)):
            return None

        try:
            return self.ws.cell(row=row_no, column=col_no).value
        except:
            traceback.print_exc()

    def get_max_col(self):
        try:
            return self.ws.max_column
        except:
            traceback.print_exc()
            return None

    def write_cell(self, row_no, col_no, content):
        """参数行号和列表从1开始表示第一行"""
        if (not isinstance(row_no, int)) or (not isinstance(col_no, int)):
            return None
        try:
            self.ws.cell(row=row_no, column=col_no).value = content
            self.wb.save(self.excel_file_path)
        except:
            traceback.print_exc()

    def write_cell_datetime(self, row_no, col_no):
        timeTup = time.localtime()
        currentDate = str(timeTup.tm_year) + "-" + \
                      str(timeTup.tm_mon) + "-" + str(timeTup.tm_mday)
        currentTime = str(timeTup.tm_hour) + ":" + \
                      str(timeTup.tm_min) + ":" + str(timeTup.tm_sec)
        self.write_cell(row_no, col_no, currentDate + " " + currentTime)

    def _save(self):
        # 表格中写入数据，保存生效
        try:
            self.wb.save(self.excel_file_path)
        except PermissionError as e:
            raise PermissionError("Please close file first.")


if __name__ == "__main__":
    wb = ParseExcel()
    print(wb.get_row(1))

